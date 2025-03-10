import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def create_financial_plan(output_filename="FinancialPlan.xlsx"):
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws_assumptions = wb.active
    ws_assumptions.title = "Assumptions"

    # ----------------------------------------------------------------
    # 1. Populate the "Assumptions" sheet
    # ----------------------------------------------------------------
    # Header row
    ws_assumptions["A1"] = "Item"
    ws_assumptions["B1"] = "Value"
    ws_assumptions["C1"] = "Comments"

    # Make headers bold
    header_font = Font(bold=True)
    ws_assumptions["A1"].font = header_font
    ws_assumptions["B1"].font = header_font
    ws_assumptions["C1"].font = header_font

    # Set column widths
    ws_assumptions.column_dimensions["A"].width = 40
    ws_assumptions.column_dimensions["B"].width = 15
    ws_assumptions.column_dimensions["C"].width = 50

    # Define assumption rows (row, item, value, comment)
    assumptions_data = [
        (2, "Time Period (months)", 6, "March 2025 - August 2025"),
        (3, "Personnel Cost (Olivier)", 10000, "€ per month"),
        (4, "Contractor (Louis)", 5000, "€ per month"),
        (5, "Overhead", 2550, "€ per month"),
        (6, "Model Usage (API)", 1000, "€ per month"),
        (7, "Cloud / Datacenter", 1000, "€ per month (Azure, HPC, GPUs)"),
        (8, "Total Fixed Costs (per month)", "=B3+B4+B5+B6+B7", "Calculated"),
        (9, "Contingency Rate (%)", 10, "10% of total fixed costs per month"),
        (10, "Contingency (per month)", "=B8*(B9/100)", "Calculated"),
        (11, "Total Monthly Outflow", "=B8+B10", "Fixed + Contingency"),
        (12, "Initial Microfinancing", 25000, "€ at Month 1"),
        (13, "Commercial Injection (Month 5)", 500000, "Min. via license or PE"),
    ]

    for row_num, item, value, comment in assumptions_data:
        ws_assumptions[f"A{row_num}"] = item
        ws_assumptions[f"B{row_num}"] = value
        ws_assumptions[f"C{row_num}"] = comment

    # ----------------------------------------------------------------
    # 2. Create and populate the "Monthly Cash Flow" sheet
    # ----------------------------------------------------------------
    ws_cf = wb.create_sheet("Monthly Cash Flow")

    # Column headers: months
    month_headers = ["Line Item", "Mar 2025 (M1)", "Apr 2025 (M2)", 
                     "May 2025 (M3)", "Jun 2025 (M4)", 
                     "Jul 2025 (M5)", "Aug 2025 (M6)", "Total"]

    for col_index, header in enumerate(month_headers, start=1):
        cell = ws_cf.cell(row=1, column=col_index, value=header)
        cell.font = Font(bold=True)

    # Set column widths
    for i in range(len(month_headers)):
        ws_cf.column_dimensions[get_column_letter(i+1)].width = 18

    # Basic structure (A) Inflows, (B) Outflows, etc.
    row_idx = 2
    ws_cf.cell(row=row_idx, column=1, value="A) Inflows").font = Font(bold=True)

    # 1. Microfinancing
    row_idx += 1
    ws_cf.cell(row=row_idx, column=1, value="1. Initial Microfinancing")
    # Put the entire microfinancing in Month 1
    ws_cf.cell(row=row_idx, column=2, value="=Assumptions!B12")  # referencing initial microfinancing
    # Zero for other months
    for col_idx in range(3, 7):  # M2 to M5
        ws_cf.cell(row=row_idx, column=col_idx, value=0)
    # We'll handle Month 6 as well
    ws_cf.cell(row=row_idx, column=7, value=0)

    # 2. Existing Resources (placeholder for any internal funding)
    row_idx += 1
    ws_cf.cell(row=row_idx, column=1, value="2. Existing Resources")
    # Put some numeric placeholder or formula here (manually or as needed).
    # For example, let's assume we place 61020 in Month 1 to cover initial shortfall:
    ws_cf.cell(row=row_idx, column=2, value=61020)  
    for col_idx in range(3, 7):
        ws_cf.cell(row=row_idx, column=col_idx, value=0)
    ws_cf.cell(row=row_idx, column=7, value=0)

    # 3. Commercial Injection (Month 5)
    row_idx += 1
    ws_cf.cell(row=row_idx, column=1, value="3. Commercial Injection (M5)")
    # Zero for months 1–4
    for col_idx in range(2, 6):
        ws_cf.cell(row=row_idx, column=col_idx, value=0)
    # Month 5 (column 6 in the sheet)
    ws_cf.cell(row=row_idx, column=6, value="=Assumptions!B13")  
    # Zero for month 6
    ws_cf.cell(row=row_idx, column=7, value=0)

    # Sum up total inflows per month
    row_idx += 1
    ws_cf.cell(row=row_idx, column=1, value="Total Inflows (A)")
    # For columns 2 to 7 (M1 to M6), sum the rows above
    for col_idx in range(2, 8):
        start_row = row_idx - 3  # The 3 lines of inflows
        end_row = row_idx - 1
        formula = f"=SUM({get_column_letter(col_idx)}{start_row}:{get_column_letter(col_idx)}{end_row})"
        ws_cf.cell(row=row_idx, column=col_idx, value=formula)

    # B) Outflows
    row_idx += 2
    ws_cf.cell(row=row_idx, column=1, value="B) Outflows").font = Font(bold=True)

    # Next lines: Personnels, Overhead, etc.
    cost_items = [
        "1. Personnel (Olivier)",
        "2. Contractor (Louis)",
        "3. Overhead",
        "4. Model Usage (API)",
        "5. Cloud / Datacenter"
    ]

    # We'll store each cost item in a row. All months reference the "Assumptions" sheet for each cost.
    # For example, "Personnel (Olivier)" = =Assumptions!B3, etc.
    cost_references = [
        "=Assumptions!B3",
        "=Assumptions!B4",
        "=Assumptions!B5",
        "=Assumptions!B6",
        "=Assumptions!B7"
    ]

    start_cost_row = row_idx + 1
    for i, (item, ref) in enumerate(zip(cost_items, cost_references)):
        row_idx += 1
        ws_cf.cell(row=row_idx, column=1, value=item)
        # For each month (2 -> 7), reference the same monthly cost
        for col_idx in range(2, 8):
            ws_cf.cell(row=row_idx, column=col_idx, value=ref)

    # Subtotal fixed costs
    row_idx += 1
    ws_cf.cell(row=row_idx, column=1, value="Subtotal Fixed Costs")
    for col_idx in range(2, 8):
        start_row = start_cost_row
        end_row = row_idx - 1
        # sum the above cost lines
        formula = f"=SUM({get_column_letter(col_idx)}{start_row}:{get_column_letter(col_idx)}{end_row})"
        ws_cf.cell(row=row_idx, column=col_idx, value=formula)

    # Contingency row
    row_idx += 1
    ws_cf.cell(row=row_idx, column=1, value="6. Contingency (10%)")
    # For each month: = Subtotal fixed costs * (Assumptions!B9 / 100)
    # Another approach: directly reference "Total Monthly Outflow" from assumptions,
    # but let's do it step by step
    for col_idx in range(2, 8):
        # row for Subtotal fixed = row_idx - 1
        sub_total_row = row_idx - 1
        formula = f"={get_column_letter(col_idx)}{sub_total_row}*(Assumptions!B9/100)"
        ws_cf.cell(row=row_idx, column=col_idx, value=formula)

    # Total Outflows (B)
    row_idx += 1
    ws_cf.cell(row=row_idx, column=1, value="Total Outflows (B)")
    for col_idx in range(2, 8):
        up_row_1 = row_idx - 2  # for contingency
        up_row_2 = row_idx - 3  # for fixed costs
        formula = f"=({get_column_letter(col_idx)}{up_row_1}+{get_column_letter(col_idx)}{up_row_2})"
        ws_cf.cell(row=row_idx, column=col_idx, value=formula)

    # C) Net Cash Flow = (A – B)
    row_idx += 2
    ws_cf.cell(row=row_idx, column=1, value="C) Net Cash Flow (A – B)").font = Font(bold=True)
    # We know "Total Inflows (A)" row and "Total Outflows (B)" row above.
    # Let's search them or track them:
    inflows_row = 6   # row index where "Total Inflows (A)" was created
    outflows_row = row_idx - 1  # row index where "Total Outflows (B)" is
    for col_idx in range(2, 8):
        formula = f"=({get_column_letter(col_idx)}{inflows_row}-{get_column_letter(col_idx)}{outflows_row})"
        ws_cf.cell(row=row_idx, column=col_idx, value=formula)

    # D) Cumulative Cash Position
    row_idx += 1
    ws_cf.cell(row=row_idx, column=1, value="D) Cumulative Cash Position").font = Font(bold=True)
    # For month 1 (col=2), it's the same as Net Cash Flow for month 1
    ws_cf.cell(row=row_idx, column=2, value=f"={get_column_letter(2)}{row_idx-1}")
    # For months 2..6, cumulative = prior cumulative + net CF current month
    for col_idx in range(3, 8):
        formula = (
            f"={get_column_letter(col_idx-1)}{row_idx}+"
            f"{get_column_letter(col_idx)}{row_idx-1}"
        )
        ws_cf.cell(row=row_idx, column=col_idx, value=formula)

    # Finally, let's add a "Total" column for some of these lines
    # For Inflows (row=6), Outflows (row=...?), Net CF (row=...), etc.
    # We can sum across columns 2..7 in the "Total" column (col=8).
    for sum_row in [4, 5, 6,  # Inflows block
                    outflows_row - 2, outflows_row - 1, outflows_row,  # Outflows block
                    row_idx - 1, row_idx]:  # Net CF and Cumulative
        cell_label = ws_cf.cell(row=sum_row, column=8)
        # sum columns B..G (2..7)
        sum_formula = f"=SUM(B{sum_row}:G{sum_row})"
        cell_label.value = sum_formula

    # Optional: format the entire sheet with an alignment (left/center) or number format
    for row in ws_cf.iter_rows(min_row=1, max_col=8):
        for cell in row:
            # Align headers center, data right or general
            if cell.row == 1:
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="right")

    # Save the workbook
    wb.save(output_filename)
    print(f"Financial plan saved as {output_filename}")

if __name__ == "__main__":
    create_financial_plan("FinancialPlan.xlsx")